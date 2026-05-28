from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


MONEY_COLUMNS = {"amount", "unit_price", "total_amount", "value"}


def _append_dataframe(ws, df: pd.DataFrame) -> None:
    # pandasのDataFrameをヘッダー付きでopenpyxlワークシートへ流し込む。
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    # 各シート共通で、ヘッダー行を見やすく装飾する。
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # 内容に合わせて列幅を調整し、金額系の列は桁区切り表示にする。
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "")
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            if header in MONEY_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def _write_key_value_rows(ws, rows: Iterable[tuple[str, object]]) -> None:
    ws.append(["項目", "内容"])
    for key, value in rows:
        ws.append([key, value])


def write_report(
    output_path: str | Path,
    input_files: list[str | Path],
    summary_metrics: pd.DataFrame,
    comments: list[str],
    monthly_summary: pd.DataFrame,
    category_summary: pd.DataFrame,
    item_summary: pd.DataFrame,
    data_check: pd.DataFrame,
    raw_data: pd.DataFrame,
) -> Path:
    """Create the Excel report workbook."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    # デフォルトシートを削除し、仕様で決めたシートだけを作成する。
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    readme = wb.create_sheet("README")
    _write_key_value_rows(
        readme,
        [
            ("ツール説明", "日本語CSV/Excel対応のEC売上レポート自動生成ツール"),
            ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("入力ファイル", ", ".join(str(Path(path)) for path in input_files)),
        ],
    )

    summary = wb.create_sheet("Summary")
    # Summaryには基本指標とコメントを同じシートに縦並びで配置する。
    summary.append(["基本指標", "値"])
    for row in summary_metrics.itertuples(index=False):
        summary.append([row.metric, row.value])
    summary.append([])
    summary.append(["自動生成コメント"])
    for comment in comments:
        summary.append([comment])

    sheets = {
        "Monthly_Summary": monthly_summary,
        "Category_Summary": category_summary,
        "Item_Summary": item_summary,
        "Data_Check": data_check,
        "Raw_Data": raw_data,
    }
    for sheet_name, dataframe in sheets.items():
        # 集計結果、チェック結果、整形済み元データをそれぞれ独立シートへ出力する。
        ws = wb.create_sheet(sheet_name)
        _append_dataframe(ws, dataframe)

    for ws in wb.worksheets:
        _style_sheet(ws)

    wb.save(output)
    return output
